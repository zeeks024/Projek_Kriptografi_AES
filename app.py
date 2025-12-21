from flask import Flask, render_template, request, jsonify, send_file
from sbox_analyzer import (get_sbox, check_bijective, check_balance, calculate_nonlinearity, 
                           calculate_sac, calculate_bic_nl, calculate_bic_sac, calculate_lap, 
                           calculate_dap, calculate_differential_uniformity, calculate_algebraic_degree,
                           calculate_transparency_order, calculate_correlation_immunity,
                           get_ddt_table, get_lat_table, 
                           encrypt_image_data, decrypt_image_data, construct_sbox_from_matrix, 
                           calculate_entropy, calculate_npcr, calculate_uaci, calculate_sensitivity_metrics,
                           get_construction_steps, AES_SBOX, SBOX_44)
from aes_cipher import AESCipher
from aes import AESInput, build_workbook, list16_to_matrix4x4_columnmajor
import base64
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)

# --- Helpers ---

def parse_sbox_input(sbox_type, custom_sbox_str):
    """Parses S-Box input from type and custom string."""
    sbox = []
    if sbox_type == 'custom' or sbox_type == 'upload':
        if not custom_sbox_str:
             return None, 'Custom S-Box string required.'
        try:
            # Handle various input formats (comma separated, space separated, hex, int)
            cleaned_str = custom_sbox_str.replace('\n', ' ').replace(',', ' ')
            parts = cleaned_str.split()
            for part in parts:
                if part.startswith('0x') or part.startswith('0X'):
                    sbox.append(int(part, 16))
                else:
                    try:
                        sbox.append(int(part))
                    except ValueError:
                         # Try hex if int fails (e.g. "A5")
                         sbox.append(int(part, 16))
            
            if len(sbox) != 256:
                return None, f'Invalid S-Box length: {len(sbox)}. Must be 256.'
            
            # Validate values are 0-255
            if any(x < 0 or x > 255 for x in sbox):
                 return None, 'S-Box values must be between 0 and 255.'
                 
            return sbox, None
        except Exception as e:
            return None, f'Error parsing custom S-Box: {str(e)}'
    else:
        sbox = get_sbox(sbox_type)
        if sbox is None:
             return None, 'Invalid S-Box type.'
        return sbox, None

@app.route('/')
def index():
    return render_template('index.html')

def get_key_from_request(req):
    key_input = req.form.get('key')
    if not key_input:
        return b'This is a key123' # Default key
    
    # Encode to bytes and pad/truncate to 16 bytes
    key_bytes = key_input.encode('utf-8')
    if len(key_bytes) < 16:
        key_bytes += b'\0' * (16 - len(key_bytes))
    else:
        key_bytes = key_bytes[:16]
    return key_bytes

def generate_excel_report(trace_data, key, sbox):
    wb = Workbook()
    ws = wb.active
    ws.title = "Encryption Trace"
    
    # Headers
    headers = ["Round", "Step", "State (Hex)", "Round Key (Hex)"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Data
    for item in trace_data:
        # Flatten state (Column-major order for AES state)
        state_flat = []
        for c in range(4):
            for r in range(4):
                state_flat.append(item['state'][r][c])
        state_hex = ' '.join([f'{x:02X}' for x in state_flat])
        
        key_hex = ''
        if item['key']:
            key_flat = []
            for c in range(4):
                for r in range(4):
                    key_flat.append(item['key'][r][c])
            key_hex = ' '.join([f'{x:02X}' for x in key_flat])
            
        ws.append([str(item['round']), item['step'], state_hex, key_hex])
        
    # Adjust widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    
    return wb

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        data = request.json
        sbox_type = data.get('type')
        custom_sbox_str = data.get('custom_sbox')
        
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error:
            return jsonify({'error': error}), 400

        # Perform Analysis
        is_bijective = check_bijective(sbox)
        balance_results = check_balance(sbox)
        nl = calculate_nonlinearity(sbox)
        sac = calculate_sac(sbox)
        bic_nl = calculate_bic_nl(sbox)
        bic_sac = calculate_bic_sac(sbox)
        lap = calculate_lap(sbox)
        dap = calculate_dap(sbox)
        du = calculate_differential_uniformity(sbox)
        ad = calculate_algebraic_degree(sbox)
        to = calculate_transparency_order(sbox)
        ci = calculate_correlation_immunity(sbox)
        
        # Format S-Box for display (Hex)
        sbox_hex = [f'{x:02X}' for x in sbox]
        
        return jsonify({
            'sbox': sbox_hex,
            'is_bijective': is_bijective,
            'balance_results': balance_results,
            'metrics': {
                'nl': nl,
                'sac': sac,
                'bic_nl': bic_nl,
                'bic_sac': bic_sac,
                'lap': lap,
                'dap': dap,
                'du': du,
                'ad': ad,
                'to': to,
                'ci': ci
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/analyze_advanced', methods=['POST'])
def analyze_advanced():
    """Returns detailed tables for visualization (DDT, LAT)."""
    try:
        data = request.json
        sbox_type = data.get('type')
        custom_sbox_str = data.get('custom_sbox')
        
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error:
            return jsonify({'error': error}), 400
        
        if not sbox:
            return jsonify({'error': 'Invalid S-Box.'}), 400
            
        ddt = get_ddt_table(sbox)
        lat = get_lat_table(sbox)
        
        return jsonify({
            'ddt': ddt,
            'lat': lat
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/avalanche_check', methods=['POST'])
def avalanche_check():
    """
    Check Avalanche Effect by flipping one bit.
    """
    try:
        sbox_type = request.form.get('type')
        custom_sbox_str = request.form.get('custom_sbox')
        # We need plain 'text_input' or hex 'input_hex'
        # Let's support 'text_input' for consistency with main encrypt, 
        # but for demo we usually start with 1 block.
        text_input = request.form.get('text_input') 
        flipped_bit_idx = int(request.form.get('flipped_bit_idx', 0))
        key_input = request.form.get('key')
        
        # Get S-Box
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error:
            return jsonify({'error': error}), 400
            
        # Get Key
        key_bytes = key_input.encode('utf-8')
        if len(key_bytes) < 16:
            key_bytes += b'\0' * (16 - len(key_bytes))
        else:
            key_bytes = key_bytes[:16]
            
        # Get Input Block
        # If text, take first 16 bytes. If <16, pad.
        data_bytes = text_input.encode('utf-8')
        block = list(data_bytes[:16])
        if len(block) < 16:
            block += [0] * (16 - len(block))
            
        # Encrypt Original
        cipher = AESCipher(key_bytes, sbox)
        original_ct = cipher.encrypt_block(block)
        
        # Flip Bit
        # flipped_bit_idx is 0..127
        byte_idx = flipped_bit_idx // 8
        bit_offset = flipped_bit_idx % 8 # 0 is LSB or MSB? Usually in Python 1<<0 is LSB.
        # Let's assume standard intuitive ordering: Bit 0 is first bit of first byte.
        # But 'bit_offset' logic depends on endianness.
        # If we want visual consistency: Byte 0, Bit 7..0.
        # Let's use simple XOR for now: 1 << (7 - bit_offset) to target specific bit in byte if MSB first view.
        # Or just 1 << bit_offset if we view it naturally.
        
        modified_block = list(block)
        # Flip bit. Let's assume MSB (bit 7) is left-most in visual.
        # flipped_bit_idx 0 -> Byte 0, Bit 7
        # flipped_bit_idx 1 -> Byte 0, Bit 6
        # ...
        # flipped_bit_idx 7 -> Byte 0, Bit 0
        shift = 7 - bit_offset
        modified_block[byte_idx] ^= (1 << shift)
        
        # Encrypt Modified
        modified_ct = cipher.encrypt_block(modified_block)
        
        # Calculate Avalanche Diff
        diff_bits = []
        changed_count = 0
        for i in range(16):
            val1 = original_ct[i]
            val2 = modified_ct[i]
            xor_val = val1 ^ val2
            
            # Check bits
            for b in range(8):
                if (xor_val >> (7-b)) & 1:
                    changed_count += 1
                    diff_bits.append(i * 8 + b) # Global bit index
                    
        percent = (changed_count / 128.0) * 100
        
        return jsonify({
            'original_hex': bytes(original_ct).hex().upper(),
            'modified_hex': bytes(modified_ct).hex().upper(),
            'changed_count': changed_count,
            'percent': percent,
            'diff_bits': diff_bits
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/encrypt', methods=['POST'])
def encrypt():
    try:
        sbox_type = request.form.get('type')
        custom_sbox_str = request.form.get('custom_sbox')
        text_input = request.form.get('text_input')
        key_input = request.form.get('key')
        image_file = request.files.get('image_file') # New field
        encryption_mode = request.form.get('encryption_mode', 'ecb') # 'ecb' or 'substitution'
        
        # Get S-Box
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error:
            return jsonify({'error': error}), 400
            
        key_format = 'text' # Force text only
        
        # Get Key Logic (Shared)
        if not key_input: key_input = "This is a key123"
        
        key_bytes = key_input.encode('utf-8')

        # Ensure key length (16, 24, 32)
        if len(key_bytes) not in [16, 24, 32]:
            if len(key_bytes) < 16: key_bytes += b'\0' * (16 - len(key_bytes))
            elif len(key_bytes) < 24: key_bytes = key_bytes[:16]
            elif len(key_bytes) < 32: key_bytes = key_bytes[:24]
            else: key_bytes = key_bytes[:32]

        cipher = AESCipher(key_bytes, sbox)

        # Handle Image Encryption
        if image_file:
            image_bytes = image_file.read()
            # Pass Raw Key + Format to Analyzer
            encrypted_b64, hist_orig, hist_enc, _ = encrypt_image_data(image_bytes, sbox, key_input, encryption_mode, key_format)
            
            if not encrypted_b64:
                 # If analyzer failed (e.g. hex error inside), it returns None
                 return jsonify({'error': 'Failed to encrypt. Check Key Format.'}), 400
                 
            # Convert original for display too
            image_file.seek(0)
            original_b64 = base64.b64encode(image_file.read()).decode('utf-8')
            
            return jsonify({
                'type': 'image',
                'original_image': f'data:image/png;base64,{original_b64}',
                'encrypted_image': f'data:image/png;base64,{encrypted_b64}',
                'hist_original': hist_orig,
                'hist_encrypted': hist_enc
            })

        elif text_input:
             # Text Encryption Logic (Legacy)
             data = text_input.encode('utf-8')
             encrypted_bytes = cipher.encrypt_data(data)
             encrypted_hex = encrypted_bytes.hex().upper()
             
             # Trace
             block = list(data[:16])
             if len(block) < 16: block += [0] * (16 - len(block))
             _, trace = cipher.encrypt_block(block, trace=True)
             
             return jsonify({
                 'encrypted_text': encrypted_hex,
                 'trace_data': trace
             })

        else:
             return jsonify({'error': 'No input provided (text or image).'}), 400
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/decrypt', methods=['POST'])
def decrypt():
    try:
        sbox_type = request.form.get('type')
        custom_sbox_str = request.form.get('custom_sbox')
        
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error:
            return jsonify({'error': error}), 400
                 
        # Decrypt
        ciphertext_hex = request.form.get('ciphertext_input')
        if not ciphertext_hex:
             return jsonify({'error': 'No ciphertext provided.'}), 400
             
        try:
            ciphertext_bytes = bytes.fromhex(ciphertext_hex)
        except ValueError:
            return jsonify({'error': 'Invalid hex string.'}), 400
            
        # Get Key
        key = get_key_from_request(request)
        cipher = AESCipher(key, sbox)
        
        decrypted_bytes = cipher.decrypt_data(ciphertext_bytes)
        
        try:
            decrypted_text = decrypted_bytes.decode('utf-8')
        except UnicodeDecodeError:
            # If not valid utf-8, return hex or base64 representation of decrypted bytes?
            # Or just return as string with replacement chars
            decrypted_text = decrypted_bytes.decode('utf-8', errors='replace')
            
        return jsonify({
            'decrypted_text': decrypted_text
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/encrypt_detailed', methods=['POST'])
def encrypt_detailed():
    try:
        sbox_type = request.form.get('type')
        custom_sbox_str = request.form.get('custom_sbox')
        
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error:
            return jsonify({'error': error}), 400
                 
        # Read input
        text_input = request.form.get('text_input')
        image_file = request.files.get('image')
        
        data_bytes = b''
        if text_input:
            data_bytes = text_input.encode('utf-8')
        elif image_file:
            data_bytes = image_file.read()
        else:
             return jsonify({'error': 'No input provided.'}), 400
        
        # Get Key
        key = get_key_from_request(request)
        
        # We need 16 bytes for matrix construction
        if len(key) < 16:
             key += b'\0' * (16 - len(key))
        key = key[:16]

        # Apply PKCS7 Padding to match standard AESCipher behavior
        # This ensures the "Detailed Report" matches the actual output
        pad_len = 16 - (len(data_bytes) % 16)
        padded_data = data_bytes + bytes([pad_len] * pad_len)
        
        # Take the first block of the PADDED data
        block_bytes = padded_data[:16]
        
        # Convert to matrix format for aes.py
        
        # Convert to matrix format for aes.py
        # list16_to_matrix4x4_columnmajor expects list of ints
        key_list = list(key)
        block_list = list(block_bytes)
        
        cipherkey_matrix = list16_to_matrix4x4_columnmajor(key_list)
        plaintext_matrix = list16_to_matrix4x4_columnmajor(block_list)
        
        # Construct AESInput
        aes_in = AESInput(cipherkey=cipherkey_matrix, plaintext=plaintext_matrix, sbox=sbox)
        
        # Generate Detailed Excel Workbook
        wb = build_workbook(aes_in, out_path=None)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='aes_detailed_trace.xlsx'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/analyze_image_sensitivity', methods=['POST'])
def analyze_image_sensitivity():
    """
    Perform deep image analysis (Entropy & NPCR).
    Requires: image, key, sbox.
    """
    try:
        sbox_type = request.form.get('type')
        custom_sbox_str = request.form.get('custom_sbox')
        image_file = request.files.get('image_file')
        key_input = request.form.get('key')
        key_format = request.form.get('key_format', 'text')
        encryption_mode = request.form.get('encryption_mode', 'ecb')

        if not image_file:
             return jsonify({'error': 'Image file required.'}), 400
             
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error: return jsonify({'error': error}), 400
        
        if not key_input: key_input = "This is a key123"
        
        # We pass key_input and key_format directly to encrypt_image_data
        # Note: We need key_bytes only if we manually perform something here not using sbox_analyzer functions
        # But wait, analyze_image_sensitivity does logic locally using sbox_analyzer functions.
        # But 'encrypt_image_data' now handles the key parsing.
        # So we don't need to parse it here? But wait...
        # 'calculate_entropy' takes ciphertext. 'calculate_npcr' takes 2 ciphertexts.
        # All good.
        
        # However, we need 'key_bytes' passed to encrypt_image_data?
        # NO, we updated encrypt_image_data to take (bytes, sbox, key_INPUT, mode, format).
        # So we pass 'key_input' string.

        # Read Original Image
        image_bytes = image_file.read()
        
        # 0. Pre-process: Resize to ensure 1-bit change persists.
        # encrypt_image_data resizes to max 256x256. We must do this BEFORE flipping a bit.
        from PIL import Image
        import numpy as np
        
        img = Image.open(io.BytesIO(image_bytes)).convert('RGB')
        
        # Resize if too large, similar to encrypt_image_data logic
        max_dim = 256
        if img.width > max_dim or img.height > max_dim:
            img.thumbnail((max_dim, max_dim))
            
        # Get processed base bytes
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        processed_image_bytes = buf.getvalue()
        
        # RESIZE IMAGE FOR SAFETY (Prevent Timeout/OOM on Deployment)
        img.thumbnail((256, 256)) # Max 256px dimension
        
        # Save processed (resized) image to bytes
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        processed_image_bytes = buf.getvalue()
        
        # 1. Encrypt Original (Processed) -> C1
        c1_b64, _, _, c1_bytes_raw = encrypt_image_data(processed_image_bytes, sbox, key_input, encryption_mode, key_format, max_dim=256)
        
        # 2. Calculate Entropy
        original_entropy = calculate_entropy(processed_image_bytes)
        encrypted_entropy = calculate_entropy(c1_bytes_raw)
        
        # 3. Modify Image (Flip 1 bit of processed image) -> P2
        # Re-open modified image from bytes to ensure consistency
        arr = np.array(Image.open(io.BytesIO(processed_image_bytes)))
        # Flip LSB of first pixel
        arr_mod = arr.copy()
        arr_mod[0,0,0] ^= 1 
        
        img_mod = Image.fromarray(arr_mod)
        buf_mod = io.BytesIO()
        img_mod.save(buf_mod, format='PNG')
        image_mod_bytes = buf_mod.getvalue()
        
        # 4. Encrypt Modified -> C2
        c2_b64, _, _, c2_bytes_raw = encrypt_image_data(image_mod_bytes, sbox, key_input, encryption_mode, key_format, max_dim=256)
        
        # 5. Calculate Metrics (NPCR & UACI) efficient pass
        npcr, uaci = calculate_sensitivity_metrics(c1_bytes_raw, c2_bytes_raw)
        
        return jsonify({
            'original_entropy': original_entropy,
            'entropy': encrypted_entropy,
            'npcr': npcr,
            'uaci': uaci
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/decrypt_image', methods=['POST'])
def decrypt_image_endpoint():
    try:
        sbox_type = request.form.get('type')
        custom_sbox_str = request.form.get('custom_sbox')
        encrypted_image_file = request.files.get('encrypted_image')
        key_input = request.form.get('key')
        key_format = request.form.get('key_format', 'text')
        encryption_mode = request.form.get('encryption_mode', 'ecb')
        
        if not encrypted_image_file:
             return jsonify({'error': 'Encrypted image required.'}), 400
             
        sbox, error = parse_sbox_input(sbox_type, custom_sbox_str)
        if error: return jsonify({'error': error}), 400
        
        if not key_input: key_input = "This is a key123"

        encrypted_bytes = encrypted_image_file.read()
        
        decrypted_b64 = decrypt_image_data(encrypted_bytes, sbox, key_input, encryption_mode, key_format)
        
        if not decrypted_b64:
             return jsonify({'error': 'Decryption failed.'}), 500
             
        return jsonify({
            'decrypted_image': f'data:image/png;base64,{decrypted_b64}'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/construct', methods=['POST'])
def construct():
    """
    Construct S-box from affine matrix.
    Expects JSON with:
    - affine_matrix: 8x8 array (list of lists)
    - c_constant: 8-element array (optional, defaults to C_AES)
    - sample_inputs: list of input values to show steps (optional, default: [0, 15, 255])
    """
    try:
        data = request.json
        affine_matrix = data.get('affine_matrix')
        c_constant = data.get('c_constant')

        sample_inputs = data.get('sample_inputs', [0, 15, 255])
        polynomial = data.get('polynomial', 0x11B) # Default to Standard AES
        
        # Parse polynomial if string
        if isinstance(polynomial, str):
            try:
                if polynomial.lower().startswith('0x'):
                    polynomial = int(polynomial, 16)
                else:
                    polynomial = int(polynomial)
            except:
                polynomial = 0x11B # Fallback
        
        if not affine_matrix:
            return jsonify({'error': 'Affine matrix is required.'}), 400
        
        # Validate matrix dimensions
        if len(affine_matrix) != 8:
            return jsonify({'error': 'Affine matrix must have 8 rows.'}), 400
        for row in affine_matrix:
            if len(row) != 8:
                return jsonify({'error': 'Each row must have 8 elements.'}), 400
            if any(val not in [0, 1] for val in row):
                return jsonify({'error': 'Matrix elements must be 0 or 1.'}), 400
        
        # Validate constant if provided
        if c_constant:
            if len(c_constant) != 8:
                return jsonify({'error': 'Constant must have 8 elements.'}), 400
            if any(val not in [0, 1] for val in c_constant):
                return jsonify({'error': 'Constant elements must be 0 or 1.'}), 400
        
        # Construct S-box
        sbox = construct_sbox_from_matrix(affine_matrix, c_constant, mod_poly=polynomial)
        
        # Test the S-box
        is_bijective = check_bijective(sbox)
        balance_results = check_balance(sbox)
        
        # Get construction steps for sample inputs
        construction_steps = []
        for x in sample_inputs:
            if 0 <= x <= 255:
                steps = get_construction_steps(x, affine_matrix, c_constant)
                construction_steps.append(steps)
        
        # Format S-box for display
        sbox_hex = [f'{x:02X}' for x in sbox]
        
        return jsonify({
            'sbox': sbox_hex,
            'is_bijective': is_bijective,
            'balance_results': balance_results,
            'construction_steps': construction_steps,
            'valid': is_bijective and all(r['is_balanced'] for r in balance_results)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/trace_input', methods=['POST'])
def trace_input():
    """
    Trace construction process for a specific input value.
    Expects JSON with:
    - input_value: integer (0-255)
    - affine_matrix: 8x8 array (list of lists)
    - c_constant: 8-element array (optional, defaults to C_AES)
    """
    try:
        data = request.json
        input_value = data.get('input_value')
        affine_matrix = data.get('affine_matrix')
        c_constant = data.get('c_constant')
        
        if input_value is None:
            return jsonify({'error': 'input_value is required.'}), 400
        
        if not isinstance(input_value, int) or input_value < 0 or input_value > 255:
            return jsonify({'error': 'input_value must be an integer between 0 and 255.'}), 400
        
        if not affine_matrix:
            return jsonify({'error': 'affine_matrix is required.'}), 400
        
        # Validate matrix dimensions
        if len(affine_matrix) != 8:
            return jsonify({'error': 'Affine matrix must have 8 rows.'}), 400
        for row in affine_matrix:
            if len(row) != 8:
                return jsonify({'error': 'Each row must have 8 elements.'}), 400
            if any(val not in [0, 1] for val in row):
                return jsonify({'error': 'Matrix elements must be 0 or 1.'}), 400
        
        # Validate constant if provided
        if c_constant:
            if len(c_constant) != 8:
                return jsonify({'error': 'Constant must have 8 elements.'}), 400
            if any(val not in [0, 1] for val in c_constant):
                return jsonify({'error': 'Constant elements must be 0 or 1.'}), 400
        
        # Get construction steps for this specific input
        steps = get_construction_steps(input_value, affine_matrix, c_constant)
        
        return jsonify({
            'success': True,
            'input_value': input_value,
            'construction_step': steps
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_sbox_excel(sbox_values):
    """
    Generate Excel file with S-Box in 16x16 grid format.
    sbox_values: list of 256 hex strings (e.g., ['00', '01', ..., 'FF'])
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "S-Box"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell_font = Font(name="Consolas", size=11)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add top-left corner cell
    ws.cell(row=1, column=1, value="")
    
    # Add column headers (0-F)
    for col in range(16):
        cell = ws.cell(row=1, column=col + 2, value=f"{col:X}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
        ws.column_dimensions[chr(66 + col)].width = 5
    
    # Add row headers and S-Box values
    for row in range(16):
        # Row header
        cell = ws.cell(row=row + 2, column=1, value=f"{row:X}0")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
        
        # S-Box values
        for col in range(16):
            idx = row * 16 + col
            value = sbox_values[idx] if isinstance(sbox_values[idx], str) else f"{sbox_values[idx]:02X}"
            cell = ws.cell(row=row + 2, column=col + 2, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            
            # Alternate row coloring
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Set first column width
    ws.column_dimensions['A'].width = 5
    
    # Add metadata sheet
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.cell(row=1, column=1, value="S-Box Information")
    ws_meta.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_meta.cell(row=3, column=1, value="Total Values:")
    ws_meta.cell(row=3, column=2, value=len(sbox_values))
    ws_meta.cell(row=4, column=1, value="Format:")
    ws_meta.cell(row=4, column=2, value="16x16 Grid (Hexadecimal)")
    ws_meta.cell(row=5, column=1, value="Generated:")
    ws_meta.cell(row=5, column=2, value="Cryptographic S-Box Analyzer")
    
    return wb

@app.route('/download_sbox_excel', methods=['POST'])
def download_sbox_excel():
    """
    Download constructed S-Box as Excel file.
    Expects JSON with 'sbox' array (256 hex values).
    """
    try:
        data = request.json
        sbox = data.get('sbox')
        
        if not sbox:
            return jsonify({'error': 'S-Box data is required.'}), 400
        
        if len(sbox) != 256:
            return jsonify({'error': f'Invalid S-Box length: {len(sbox)}. Must be 256.'}), 400
        
        # Generate Excel file
        wb = generate_sbox_excel(sbox)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='sbox_constructed.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download_template_sbox', methods=['GET'])
def download_template_sbox():
    """
    Download empty S-Box Excel template.
    """
    try:
        # Create 256 empty strings for the template
        empty_sbox = [""] * 256
        
        # Generate Excel file
        wb = generate_sbox_excel(empty_sbox)
        
        # Manual override: Set 16x16 grid cells to explicit empty strings if needed, 
        # but generate_sbox_excel handles it if we pass ""
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='sbox_template_empty.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def parse_sbox_excel(file_stream):
    """
    Parse Excel file and extract S-Box values from 16x16 grid.
    Expected format: Row headers in column A (00, 10, 20, ..., F0)
                    Column headers in row 1 (0, 1, 2, ..., F)
                    S-Box values in cells B2:Q17
    Returns: list of 256 integer values (0-255)
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(file_stream)
    ws = wb.active
    
    sbox = []
    
    # Read 16x16 grid starting from B2 (row 2, col 2)
    for row in range(16):
        for col in range(16):
            cell = ws.cell(row=row + 2, column=col + 2)
            value = cell.value
            
            if value is None:
                raise ValueError(f"Empty cell at row {row}, column {col}")
            
            # Convert to integer (handle hex strings or integers)
            if isinstance(value, str):
                # Remove any whitespace
                value = value.strip()
                # Try to parse as hex
                try:
                    int_value = int(value, 16)
                except ValueError:
                    # Try as decimal
                    try:
                        int_value = int(value)
                    except ValueError:
                        raise ValueError(f"Invalid value '{value}' at row {row}, column {col}")
            elif isinstance(value, (int, float)):
                int_value = int(value)
            else:
                raise ValueError(f"Unexpected value type at row {row}, column {col}")
            
            # Validate range
            if int_value < 0 or int_value > 255:
                raise ValueError(f"Value {int_value} out of range (0-255) at row {row}, column {col}")
            
            sbox.append(int_value)
    
    if len(sbox) != 256:
        raise ValueError(f"Invalid S-Box size: {len(sbox)}. Expected 256 values.")
    
    return sbox

@app.route('/upload_sbox_excel', methods=['POST'])
def upload_sbox_excel():
    """
    Upload Excel file containing S-Box and return parsed values.
    Expects Excel file in 16x16 grid format.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded.'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected.'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls).'}), 400
        
        # Parse the Excel file
        sbox = parse_sbox_excel(file)
        
        # Format as hex for frontend (prefixed with 0x to ensure it's treated as hex)
        sbox_hex = [f'0x{x:02X}' for x in sbox]
        
        return jsonify({
            'sbox': sbox_hex,
            'message': 'S-Box uploaded successfully!'
        })
        
    except ValueError as e:
        return jsonify({'error': f'Parsing error: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error reading file: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=False, port=5001)

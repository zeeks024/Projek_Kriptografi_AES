#!/usr/bin/env python3
"""
AES-128 Encryption → Excel Detail Logger
-----------------------------------------
Generates a comprehensive Excel workbook documenting complete AES-128 encryption
process for Rounds 1-10, with EVERY transformation step shown in detail:

Round 0:
- Plaintext XOR CipherKey (initial whitening)

Rounds 1-9:
- SubBytes (with S-BOX lookup details)
- ShiftRows (row rotation pattern)
- MixColumns (Galois Field multiplication details)
- AddRoundKey (XOR with round key)

Round 10:
- SubBytes
- ShiftRows
- AddRoundKey (NO MixColumns)

Features:
- Standard AES-128 (FIPS-197 compliant)
- Column-major matrix format
- Byte-by-byte transformation details
- Round key schedule visualization
- Complete SubBytes, ShiftRows, MixColumns, AddRoundKey details

Usage (CLI):
    python aes.py --plaintext-hex "32 43 F6 A8 88 5A 30 8D 31 31 98 A2 E0 37 07 34" \
                  --cipherkey-hex "2B 7E 15 16 28 AE D2 A6 AB F7 15 88 09 CF 4F 3C" \
                  --out "AES_Encryption_Detail.xlsx"

Input format: 16 hex bytes in COLUMN-MAJOR order (AES standard)
Output: Single Excel sheet "AES ENCRYPTION" with complete round details

Implementation: Standard FIPS-197 AES-128 with column-major matrix representation
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

# =======================
# AES Core (FIPS-197 Standard)
# =======================

# Standard S-BOX from FIPS-197
SBOX = [
0x63,0x7c,0x77,0x7b,0xf2,0x6b,0x6f,0xc5,0x30,0x01,0x67,0x2b,0xfe,0xd7,0xab,0x76,
0xca,0x82,0xc9,0x7d,0xfa,0x59,0x47,0xf0,0xad,0xd4,0xa2,0xaf,0x9c,0xa4,0x72,0xc0,
0xb7,0xfd,0x93,0x26,0x36,0x3f,0xf7,0xcc,0x34,0xa5,0xe5,0xf1,0x71,0xd8,0x31,0x15,
0x04,0xc7,0x23,0xc3,0x18,0x96,0x05,0x9a,0x07,0x12,0x80,0xe2,0xeb,0x27,0xb2,0x75,
0x09,0x83,0x2c,0x1a,0x1b,0x6e,0x5a,0xa0,0x52,0x3b,0xd6,0xb3,0x29,0xe3,0x2f,0x84,
0x53,0xd1,0x00,0xed,0x20,0xfc,0xb1,0x5b,0x6a,0xcb,0xbe,0x39,0x4a,0x4c,0x58,0xcf,
0xd0,0xef,0xaa,0xfb,0x43,0x4d,0x33,0x85,0x45,0xf9,0x02,0x7f,0x50,0x3c,0x9f,0xa8,
0x51,0xa3,0x40,0x8f,0x92,0x9d,0x38,0xf5,0xbc,0xb6,0xda,0x21,0x10,0xff,0xf3,0xd2,
0xcd,0x0c,0x13,0xec,0x5f,0x97,0x44,0x17,0xc4,0xa7,0x7e,0x3d,0x64,0x5d,0x19,0x73,
0x60,0x81,0x4f,0xdc,0x22,0x2a,0x90,0x88,0x46,0xee,0xb8,0x14,0xde,0x5e,0x0b,0xdb,
0xe0,0x32,0x3a,0x0a,0x49,0x06,0x24,0x5c,0xc2,0xd3,0xac,0x62,0x91,0x95,0xe4,0x79,
0xe7,0xc8,0x37,0x6d,0x8d,0xd5,0x4e,0xa9,0x6c,0x56,0xf4,0xea,0x65,0x7a,0xae,0x08,
0xba,0x78,0x25,0x2e,0x1c,0xa6,0xb4,0xc6,0xe8,0xdd,0x74,0x1f,0x4b,0xbd,0x8b,0x8a,
0x70,0x3e,0xb5,0x66,0x48,0x03,0xf6,0x0e,0x61,0x35,0x57,0xb9,0x86,0xc1,0x1d,0x9e,
0xe1,0xf8,0x98,0x11,0x69,0xd9,0x8e,0x94,0x9b,0x1e,0x87,0xe9,0xce,0x55,0x28,0xdf,
0x8c,0xa1,0x89,0x0d,0xbf,0xe6,0x42,0x68,0x41,0x99,0x2d,0x0f,0xb0,0x54,0xbb,0x16
]

RCON = [0x00,0x01,0x02,0x04,0x08,0x10,0x20,0x40,0x80,0x1B,0x36]

def _xtime(a: int) -> int:
    """Multiply by x in GF(2^8)"""
    a &= 0xFF
    res = (a << 1) & 0xFF
    if a & 0x80:
        res ^= 0x1b
    return res

def _mul(a: int, b: int) -> int:
    """Multiply in GF(2^8) - only supports b=1,2,3"""
    if b == 1: return a & 0xFF
    if b == 2: return _xtime(a)
    if b == 3: return _xtime(a) ^ (a & 0xFF)
    raise ValueError("GF multiplier must be 1, 2, or 3")

def byte_to_binary(b: int) -> str:
    """Convert byte to 8-bit binary string"""
    return f"{b:08b}"

def byte_to_polynomial(b: int) -> str:
    """Convert byte to polynomial representation"""
    if b == 0:
        return "0"
    
    terms = []
    powers = ["x⁷", "x⁶", "x⁵", "x⁴", "x³", "x²", "x", "1"]
    
    for i in range(8):
        if b & (1 << (7 - i)):
            terms.append(powers[i])
    
    return "+".join(terms) if terms else "0"

def _mul_with_detail(a: int, b: int) -> Dict:
    """Multiply in GF(2^8) with detailed steps"""
    result = _mul(a, b)
    
    detail = {
        "input_hex": f"{a:02x}",
        "input_binary": byte_to_binary(a),
        "input_poly": byte_to_polynomial(a),
        "multiplier": b,
        "multiplier_poly": byte_to_polynomial(b),
        "result_hex": f"{result:02x}",
        "result_binary": byte_to_binary(result),
        "result_poly": byte_to_polynomial(result),
    }
    
    # Add modulo info if reduction happened
    if b == 2 or b == 3:
        shifted = (a << 1) & 0x1FF  # Allow 9 bits
        if a & 0x80:  # If overflow
            detail["needs_modulo"] = True
            detail["shifted_binary"] = f"{shifted:09b}"
            detail["modulo_poly"] = "x⁴+x³+x+1"
            detail["modulo_binary"] = "100011011"
        else:
            detail["needs_modulo"] = False
    
    return detail

def sub_bytes(state: List[List[int]], sbox: List[int] = SBOX) -> List[List[int]]:
    """Apply S-BOX substitution to each byte"""
    return [[sbox[b] for b in row] for row in state]

def shift_rows(state: List[List[int]]) -> List[List[int]]:
    """Shift rows: row i shifted left by i positions"""
    return [
        state[0][:],
        state[1][1:] + state[1][:1],
        state[2][2:] + state[2][:2],
        state[3][3:] + state[3][:3],
    ]

def mix_single_column_with_detail(col: List[int], col_idx: int) -> tuple:
    """Mix single column with detailed steps for Excel including polynomial representation"""
    a0, a1, a2, a3 = col
    details = []
    
    # Row 0: 02*a0 ⊕ 03*a1 ⊕ 01*a2 ⊕ 01*a3
    r0_parts = [_mul(a0,2), _mul(a1,3), _mul(a2,1), _mul(a3,1)]
    r0 = r0_parts[0] ^ r0_parts[1] ^ r0_parts[2] ^ r0_parts[3]
    poly_details_r0 = [
        _mul_with_detail(a0, 2),
        _mul_with_detail(a1, 3),
        _mul_with_detail(a2, 1),
        _mul_with_detail(a3, 1)
    ]
    details.append({
        "out_row": 0, "col": col_idx, "input": [a0, a1, a2, a3],
        "formula": "02*a0 ⊕ 03*a1 ⊕ 01*a2 ⊕ 01*a3",
        "parts_hex": [f"{p:02x}" for p in r0_parts],
        "result_hex": f"{r0:02x}",
        "poly_details": poly_details_r0
    })
    
    # Row 1: 01*a0 ⊕ 02*a1 ⊕ 03*a2 ⊕ 01*a3
    r1_parts = [_mul(a0,1), _mul(a1,2), _mul(a2,3), _mul(a3,1)]
    r1 = r1_parts[0] ^ r1_parts[1] ^ r1_parts[2] ^ r1_parts[3]
    poly_details_r1 = [
        _mul_with_detail(a0, 1),
        _mul_with_detail(a1, 2),
        _mul_with_detail(a2, 3),
        _mul_with_detail(a3, 1)
    ]
    details.append({
        "out_row": 1, "col": col_idx, "input": [a0, a1, a2, a3],
        "formula": "01*a0 ⊕ 02*a1 ⊕ 03*a2 ⊕ 01*a3",
        "parts_hex": [f"{p:02x}" for p in r1_parts],
        "result_hex": f"{r1:02x}",
        "poly_details": poly_details_r1
    })
    
    # Row 2: 01*a0 ⊕ 01*a1 ⊕ 02*a2 ⊕ 03*a3
    r2_parts = [_mul(a0,1), _mul(a1,1), _mul(a2,2), _mul(a3,3)]
    r2 = r2_parts[0] ^ r2_parts[1] ^ r2_parts[2] ^ r2_parts[3]
    poly_details_r2 = [
        _mul_with_detail(a0, 1),
        _mul_with_detail(a1, 1),
        _mul_with_detail(a2, 2),
        _mul_with_detail(a3, 3)
    ]
    details.append({
        "out_row": 2, "col": col_idx, "input": [a0, a1, a2, a3],
        "formula": "01*a0 ⊕ 01*a1 ⊕ 02*a2 ⊕ 03*a3",
        "parts_hex": [f"{p:02x}" for p in r2_parts],
        "result_hex": f"{r2:02x}",
        "poly_details": poly_details_r2
    })
    
    # Row 3: 03*a0 ⊕ 01*a1 ⊕ 01*a2 ⊕ 02*a3
    r3_parts = [_mul(a0,3), _mul(a1,1), _mul(a2,1), _mul(a3,2)]
    r3 = r3_parts[0] ^ r3_parts[1] ^ r3_parts[2] ^ r3_parts[3]
    poly_details_r3 = [
        _mul_with_detail(a0, 3),
        _mul_with_detail(a1, 1),
        _mul_with_detail(a2, 1),
        _mul_with_detail(a3, 2)
    ]
    details.append({
        "out_row": 3, "col": col_idx, "input": [a0, a1, a2, a3],
        "formula": "03*a0 ⊕ 01*a1 ⊕ 01*a2 ⊕ 02*a3",
        "parts_hex": [f"{p:02x}" for p in r3_parts],
        "result_hex": f"{r3:02x}",
        "poly_details": poly_details_r3
    })
    
    return [r0, r1, r2, r3], details

def mix_columns_with_detail(state: List[List[int]]) -> tuple:
    """Mix columns with details for Excel"""
    cols = [[state[r][c] for r in range(4)] for c in range(4)]
    mixed = []
    all_details = []
    
    for col_idx, col in enumerate(cols):
        result_col, col_details = mix_single_column_with_detail(col, col_idx)
        mixed.append(result_col)
        all_details.extend(col_details)
    
    mixed_state = [[mixed[c][r] for c in range(4)] for r in range(4)]
    return mixed_state, all_details

def add_round_key(state: List[List[int]], rk: List[List[int]]) -> List[List[int]]:
    """XOR state with round key"""
    return [[state[i][j] ^ rk[i][j] for j in range(4)] for i in range(4)]

def key_schedule_from_cipherkey_matrix(cipherkey: List[List[int]], sbox: List[int] = SBOX) -> List[List[List[int]]]:
    """
    Generate AES-128 round keys using standard FIPS-197 key expansion algorithm.
    
    Input: cipherkey as 4×4 matrix (column-major representation)
    Output: 11 round keys [RK0..RK10], each as 4×4 matrix
    
    Algorithm (Standard AES-128 Key Expansion):
    - RK0 = CipherKey (initial key, no transformation)
    - For rounds 1-10:
      - RKn[*][0] = RK(n-1)[*][0] ⊕ g(RK(n-1)[*][3], Rcon[n])
      - RKn[*][j] = RK(n-1)[*][j] ⊕ RKn[*][j-1] for j=1,2,3
    
    where g(col, rcon) = SubWord(RotWord(col)) ⊕ [rcon, 0, 0, 0]
    - RotWord: rotate column bytes [a,b,c,d] → [b,c,d,a]
    - SubWord: apply S-BOX to each byte
    - Rcon: round constant (only applied to first byte)
    """
    rks = [cipherkey]  # RK0 = CipherKey
    
    for round_num in range(1, 11):
        prev_rk = rks[round_num - 1]
        new_rk = [[0]*4 for _ in range(4)]
        
        # g-function on last column of previous RK
        last_col = [prev_rk[r][3] for r in range(4)]
        g = last_col[1:] + last_col[:1]  # RotWord
        g = [sbox[x] for x in g]  # SubWord
        g[0] ^= RCON[round_num]  # XOR Rcon
        
        # Column 0: prev_rk col 0 ⊕ g
        for r in range(4):
            new_rk[r][0] = prev_rk[r][0] ^ g[r]
        
        # Columns 1-3: prev_rk col j ⊕ new_rk col j-1
        for col in range(1, 4):
            for r in range(4):
                new_rk[r][col] = prev_rk[r][col] ^ new_rk[r][col-1]
        
        rks.append(new_rk)
    
    return rks

# =======================
# Excel Utilities
# =======================

THIN = Side(style="thin", color="000000")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="F2F2F2")

def hex2(v: int) -> str:
    """Convert byte to 2-digit hex string"""
    return f"{v:02x}"

def write_block(ws: Worksheet, r0: int, c0: int, block: List[List[int]], title: Optional[str]=None) -> None:
    """Write 4x4 matrix to Excel with formatting"""
    if title:
        ws.cell(row=r0-1, column=c0, value=title).font = Font(bold=True)
        ws.cell(row=r0-1, column=c0).fill = HDR_FILL
    for i in range(4):
        for j in range(4):
            cell = ws.cell(row=r0+i, column=c0+j, value=hex2(block[i][j]))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Consolas")
            cell.border = BOX_BORDER

def freeze_at(ws: Worksheet, cell: str):
    """Freeze panes at specified cell"""
    ws.freeze_panes = cell

# =======================
# Pipeline
# =======================

@dataclass
class AESInput:
    cipherkey: List[List[int]]  # 4×4 matrix
    plaintext: List[List[int]]  # 4×4 matrix
    sbox: List[int] = None # Optional custom S-Box

def compute_all_rounds(aes_in: AESInput) -> Tuple[List[List[List[int]]], List[Dict]]:
    """Compute all AES rounds with detailed history"""
    current_sbox = aes_in.sbox if aes_in.sbox else SBOX
    rks = key_schedule_from_cipherkey_matrix(aes_in.cipherkey, sbox=current_sbox)
    states_history = []
    
    # ROUND 0: Plaintext XOR CipherKey (initial whitening)
    prev = add_round_key(aes_in.plaintext, aes_in.cipherkey)
    
    # ROUND 1-10
    for r in range(1, 11):
        # Round 10: NO MixColumns
        if r == 10:
            s_sub = sub_bytes(prev, sbox=current_sbox)
            s_shift = shift_rows(s_sub)
            s_mix = s_shift
            mixcolumns_detail = []
        else:
            s_sub = sub_bytes(prev, sbox=current_sbox)
            s_shift = shift_rows(s_sub)
            s_mix, mixcolumns_detail = mix_columns_with_detail(s_shift)
        
        s_out = add_round_key(s_mix, rks[r])
        
        # SubBytes details
        subbytes_detail = []
        for i in range(4):
            for j in range(4):
                subbytes_detail.append({
                    "row": i, "col": j,
                    "input_hex": f"{prev[i][j]:02x}",
                    "input_dec": prev[i][j],
                    "sbox_output_hex": f"{s_sub[i][j]:02x}",
                    "sbox_output_dec": s_sub[i][j]
                })
        
        # AddRoundKey details
        addroundkey_detail = []
        for i in range(4):
            for j in range(4):
                addroundkey_detail.append({
                    "row": i, "col": j,
                    "state_hex": f"{s_mix[i][j]:02x}",
                    "state_dec": s_mix[i][j],
                    "key_hex": f"{rks[r][i][j]:02x}",
                    "key_dec": rks[r][i][j],
                    "xor_result_hex": f"{s_out[i][j]:02x}",
                    "xor_result_dec": s_out[i][j],
                    "xor_binary": f"{s_mix[i][j]:08b} ⊕ {rks[r][i][j]:08b} = {s_out[i][j]:08b}"
                })
        
        states_history.append({
            "round": r,
            "input": prev,
            "subbytes": s_sub,
            "subbytes_detail": subbytes_detail,
            "shiftrows": s_shift,
            "mix_or_nomix": s_mix,
            "mixcolumns_detail": mixcolumns_detail,
            "after_addroundkey": s_out,
            "addroundkey_detail": addroundkey_detail,
            "roundkey": rks[r],
        })
        prev = s_out
    
    return rks, states_history

def sheet_all_in_one(wb: Workbook, cipherkey: List[List[int]], plaintext: List[List[int]], history: List[dict]) -> Worksheet:
    """Create comprehensive AES encryption detail sheet"""
    ws = wb.create_sheet("AES ENCRYPTION")
    
    # Set column widths
    widths = {1: 3, 2: 15, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 22, 11: 3}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    freeze_at(ws, "A4")
    row = 2
    
    # HEADER UTAMA
    ws.merge_cells(f"B{row}:J{row}")
    cell = ws.cell(row=row, column=2, value="AES ENCRYPTION")
    cell.font = Font(bold=True, size=18, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    # INPUT INFO
    ws.merge_cells(f"B{row}:J{row}")
    ws.cell(row=row, column=2, value="📋 INFORMASI INPUT").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="4472C4")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    row += 2
    
    ws.cell(row=row, column=2, value="Plaintext (Input Asli):").font = Font(bold=True)
    row += 1
    write_block(ws, row, 2, plaintext)
    row += 6
    
    ws.cell(row=row, column=2, value="CipherKey (Kunci/Master Key):").font = Font(bold=True)
    row += 1
    write_block(ws, row, 2, cipherkey)
    row += 6
    
    # PROCESS EACH ROUND
    for h in history:
        r = h["round"]
        
        # ROUND HEADER
        ws.merge_cells(f"A{row}:K{row}")
        cell = ws.cell(row=row, column=1, value=f"═══════════════ ROUND {r} ═══════════════")
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="203864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        # INPUT STATE
        ws.merge_cells(f"B{row}:J{row}")
        input_label = "Plaintext" if r == 1 else f"Output Round {r-1}"
        ws.cell(row=row, column=2, value=f"📥 INPUT STATE Round {r} (= {input_label})").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="7030A0")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1
        
        write_block(ws, row, 2, h["input"], title="Input Matrix (Hex)")
        row += 6
        
        # STEP 1: SUBBYTES
        ws.merge_cells(f"B{row}:J{row}")
        ws.cell(row=row, column=2, value="🔄 STEP 1: SubBytes Transformation (S-Box Lookup)").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="70AD47")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1
        
        # SubBytes table header
        headers = ["Posisi", "Input Hex", "Input Dec", "S-Box →", "Output Hex", "Output Dec"]
        for col_offset, header in enumerate(headers):
            cell = ws.cell(row=row, column=2+col_offset, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BOX_BORDER
        row += 1
        
        # SubBytes data
        for detail in h["subbytes_detail"]:
            ws.cell(row=row, column=2, value=f"[{detail['row']},{detail['col']}]").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=detail["input_hex"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=detail["input_dec"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value="→").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=6, value=detail["sbox_output_hex"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=7, value=detail["sbox_output_dec"]).alignment = Alignment(horizontal="center")
            
            for col in range(2, 8):
                ws.cell(row=row, column=col).border = BOX_BORDER
                ws.cell(row=row, column=col).font = Font(name="Consolas", size=9)
            row += 1
        
        row += 1
        write_block(ws, row, 2, h["subbytes"], title="✓ Hasil SubBytes Matrix")
        row += 6
        
        # STEP 2: SHIFTROWS
        ws.merge_cells(f"B{row}:J{row}")
        ws.cell(row=row, column=2, value="🔄 STEP 2: ShiftRows Transformation").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FFC000")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1
        
        ws.cell(row=row, column=2, value="• Row 0: Tidak dishift (0 posisi)").font = Font(italic=True)
        row += 1
        ws.cell(row=row, column=2, value="• Row 1: Shift kiri 1 posisi").font = Font(italic=True)
        row += 1
        ws.cell(row=row, column=2, value="• Row 2: Shift kiri 2 posisi").font = Font(italic=True)
        row += 1
        ws.cell(row=row, column=2, value="• Row 3: Shift kiri 3 posisi").font = Font(italic=True)
        row += 2
        
        write_block(ws, row, 2, h["shiftrows"], title="✓ Hasil ShiftRows Matrix")
        row += 6
        
        # STEP 3: MIXCOLUMNS
        if r != 10:
            ws.merge_cells(f"B{row}:J{row}")
            ws.cell(row=row, column=2, value="🔄 STEP 3: MixColumns Transformation (Galois Field GF(2⁸))").font = Font(bold=True, size=11, color="FFFFFF")
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="5B9BD5")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            row += 1
            
            ws.cell(row=row, column=2, value="Matrix: [[02,03,01,01], [01,02,03,01], [01,01,02,03], [03,01,01,02]]").font = Font(italic=True, size=9)
            row += 1
            
            # Polinom info
            ws.cell(row=row, column=2, value="Polinom pada GF(2⁸): x⁴+x³+x+1 (digunakan untuk modulo reduksi pada perkalian)").font = Font(italic=True, size=9, color="0066CC")
            row += 1
            
            # MixColumns table header
            headers_mix = ["Output Pos", "Column", "Formula", "Input", "Multiply", "XOR", "Result"]
            for col_offset, header in enumerate(headers_mix):
                cell = ws.cell(row=row, column=2+col_offset, value=header)
                cell.font = Font(bold=True, size=9)
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center")
                cell.border = BOX_BORDER
            row += 1
            
            # MixColumns data
            for detail in h["mixcolumns_detail"]:
                ws.cell(row=row, column=2, value=f"[{detail['out_row']},{detail['col']}]").alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=3, value=f"Col{detail['col']}").alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4, value=detail["formula"]).alignment = Alignment(horizontal="left")
                
                input_str = " ".join([f"{b:02x}" for b in detail["input"]])
                ws.cell(row=row, column=5, value=input_str).alignment = Alignment(horizontal="center")
                
                parts_str = " ⊕ ".join(detail["parts_hex"])
                ws.cell(row=row, column=6, value=parts_str).alignment = Alignment(horizontal="center")
                
                # XOR result - menggunakan teks biasa tanpa tanda "="
                ws.cell(row=row, column=7, value=detail["result_hex"]).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=8, value=detail["result_hex"]).alignment = Alignment(horizontal="center")
                
                for col in range(2, 9):
                    ws.cell(row=row, column=col).border = BOX_BORDER
                    ws.cell(row=row, column=col).font = Font(name="Consolas", size=9)
                row += 1
            
            row += 1
        else:
            ws.merge_cells(f"B{row}:J{row}")
            ws.cell(row=row, column=2, value="⊗ STEP 3: MixColumns DILEWATI (Round 10 tidak menggunakan MixColumns)").font = Font(bold=True, size=11, color="FFFFFF")
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="C00000")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            row += 2
        
        write_block(ws, row, 2, h["mix_or_nomix"], title="✓ Hasil MixColumns Matrix" if r!=10 else "✓ State Matrix (sama dengan ShiftRows)")
        row += 6
        
        # STEP 4: ADDROUNDKEY
        ws.merge_cells(f"B{row}:J{row}")
        ws.cell(row=row, column=2, value="🔄 STEP 4: AddRoundKey (XOR dengan Round Key)").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="70AD47")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1
        
        # AddRoundKey table header
        headers_ark = ["Posisi", "State Hex", "State Dec", "Key Hex", "Key Dec", "⊕", "Result Hex", "Result Dec", "Binary Operation"]
        for col_offset, header in enumerate(headers_ark):
            cell = ws.cell(row=row, column=2+col_offset, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BOX_BORDER
        row += 1
        
        # AddRoundKey data
        for detail in h["addroundkey_detail"]:
            ws.cell(row=row, column=2, value=f"[{detail['row']},{detail['col']}]").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=detail["state_hex"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=detail["state_dec"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=detail["key_hex"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=6, value=detail["key_dec"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=7, value="⊕").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=8, value=detail["xor_result_hex"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=9, value=detail["xor_result_dec"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=10, value=detail["xor_binary"]).alignment = Alignment(horizontal="left")
            
            for col in range(2, 11):
                ws.cell(row=row, column=col).border = BOX_BORDER
                ws.cell(row=row, column=col).font = Font(name="Consolas", size=8)
            row += 1
        
        row += 1
        
        # Round Key
        write_block(ws, row, 2, h["roundkey"], title=f"Round Key {r} yang digunakan")
        row += 6
        
        # OUTPUT
        ws.merge_cells(f"B{row}:J{row}")
        ws.cell(row=row, column=2, value=f"✅ OUTPUT ROUND {r} (State setelah AddRoundKey)").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="375623")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1
        
        write_block(ws, row, 2, h["after_addroundkey"])
        row += 6
        
        # Separator
        ws.merge_cells(f"A{row}:K{row}")
        ws.cell(row=row, column=1, value="").fill = PatternFill("solid", fgColor="D9D9D9")
        row += 3
    
    # FINAL RESULT
    ws.merge_cells(f"B{row}:J{row}")
    cell = ws.cell(row=row, column=2, value="🎯 HASIL AKHIR ENKRIPSI AES")
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="C00000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    final = history[-1]["after_addroundkey"]
    ws.cell(row=row, column=2, value="Ciphertext (Output Round 10):").font = Font(bold=True, size=11)
    row += 1
    write_block(ws, row, 2, final)
    
    return ws

def sheet_polynomial_details(wb: Workbook, history: List[Dict]) -> Worksheet:
    """Create detailed polynomial representation sheet for MixColumns"""
    ws = wb.create_sheet("Detail Polinom MixColumns")
    
    # Title
    ws.merge_cells("B2:K2")
    cell = ws.cell(2, 2, "DETAIL PERHITUNGAN POLINOM - MIXCOLUMNS")
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row = 4
    ws.cell(row, 2, "Penjelasan:").font = Font(bold=True, size=11)
    row += 1
    ws.cell(row, 2, "• Setiap byte direpresentasikan dalam Binary dan Polynomial")
    row += 1
    ws.cell(row, 2, "• Perkalian di GF(2⁸) menggunakan modulo x⁸+x⁴+x³+x+1 (0x1b dalam hex)")
    row += 1
    ws.cell(row, 2, "• Jika hasil perkalian overflow (bit ke-8 set), dilakukan XOR dengan 0x1b")
    row += 3
    
    # Process Round 1 as example
    for h in history:
        r = h["round"]
        if r != 1:  # Hanya tampilkan Round 1 sebagai contoh
            continue
        
        ws.merge_cells(f"B{row}:K{row}")
        cell = ws.cell(row, 2, f"═══════════ ROUND {r} - CONTOH DETAIL POLINOM ═══════════")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="203864")
        cell.alignment = Alignment(horizontal="center")
        row += 2
        
        # Show first calculation as example
        detail = h["mixcolumns_detail"][0]  # [0,0]
        
        ws.cell(row, 2, f"Perhitungan untuk posisi [{detail['out_row']},{detail['col']}]").font = Font(bold=True, size=10, color="0066CC")
        row += 1
        ws.cell(row, 2, f"Formula: {detail['formula']}")
        row += 1
        ws.cell(row, 2, f"Input bytes: {' '.join(f'{b:02x}' for b in detail['input'])}")
        row += 2
        
        # Header
        headers = ["Byte", "Hex", "Binary (8-bit)", "Polynomial", "×", "Mult", "→", "Result Hex", "Result Bin", "Result Poly", "Modulo"]
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row, 2 + col_idx, header)
            cell.font = Font(bold=True, size=9)
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BOX_BORDER
        row += 1
        
        # Show details
        byte_names = ["a0 (d4)", "a1 (bf)", "a2 (5d)", "a3 (30)"]
        multipliers = [2, 3, 1, 1]
        for idx, poly_det in enumerate(detail["poly_details"]):
            ws.cell(row, 2, byte_names[idx]).font = Font(bold=True)
            ws.cell(row, 3, poly_det["input_hex"])
            ws.cell(row, 4, poly_det["input_binary"]).font = Font(name="Consolas", size=8)
            ws.cell(row, 5, poly_det["input_poly"]).font = Font(size=8)
            ws.cell(row, 6, "×").alignment = Alignment(horizontal="center")
            ws.cell(row, 7, f"{poly_det['multiplier']} ({poly_det['multiplier_poly']})")
            ws.cell(row, 8, "=").alignment = Alignment(horizontal="center")
            ws.cell(row, 9, poly_det["result_hex"]).font = Font(bold=True, color="0066CC")
            ws.cell(row, 10, poly_det["result_binary"]).font = Font(name="Consolas", size=8)
            ws.cell(row, 11, poly_det["result_poly"]).font = Font(size=8)
            
            if poly_det.get("needs_modulo"):
                cell = ws.cell(row, 12, "✓ XOR 0x1b")
                cell.font = Font(color="FF0000", bold=True)
                cell.fill = PatternFill("solid", fgColor="FFE6E6")
            else:
                ws.cell(row, 12, "-")
            
            row += 1
        
        row += 1
        ws.cell(row, 2, "XOR semua hasil:").font = Font(bold=True)
        row += 1
        xor_formula = f"{detail['parts_hex'][0]} ⊕ {detail['parts_hex'][1]} ⊕ {detail['parts_hex'][2]} ⊕ {detail['parts_hex'][3]} = {detail['result_hex']}"
        ws.cell(row, 2, xor_formula).font = Font(bold=True, size=11, color="0066CC")
        row += 1
        result_int = int(detail['result_hex'], 16)
        ws.cell(row, 2, f"Binary: {byte_to_binary(result_int)}").font = Font(italic=True)
        row += 1
        ws.cell(row, 2, f"Polynomial: {byte_to_polynomial(result_int)}").font = Font(italic=True, color="0066CC")
        row += 3
        
        break
    
    # Set column widths
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 12
    
    return ws

def build_workbook(aes_in: AESInput, out_path: Optional[Path] = None) -> Union[Path, Workbook]:
    """Build Excel workbook with AES encryption details. Returns Path if saved, or Workbook object if out_path is None."""
    rks, history = compute_all_rounds(aes_in)
    wb = Workbook()
    del wb[wb.sheetnames[0]]  # Remove default sheet
    
    sheet_all_in_one(wb, aes_in.cipherkey, aes_in.plaintext, history)
    sheet_polynomial_details(wb, history)
    
    if out_path:
        wb.save(out_path)
        print(f"\n✅ Excel file berhasil dibuat dengan 2 sheets!")
        print(f"   Sheet 1: {wb.sheetnames[0]} (Main AES encryption process)")
        print(f"   Sheet 2: {wb.sheetnames[1]} (Polynomial details for MixColumns)")
        print(f"   Total rounds: 11 rounds (Round 0-10)")
        print(f"   Round 0: Plaintext XOR CipherKey (initial whitening)")
        print(f"   Round 1-9: SubBytes + ShiftRows + MixColumns + AddRoundKey")
        print(f"   Round 10: SubBytes + ShiftRows + AddRoundKey (tanpa MixColumns)")
        return out_path
    
    return wb

# =======================
# CLI & I/O
# =======================

def read_from_source_excel(src_path: Path) -> AESInput:
    """Read CipherKey and Plaintext from Excel file"""
    wb = load_workbook(src_path, data_only=False)
    ws = wb["Sheet1"]
    
    def read_block(tl: str, br: str) -> List[List[int]]:
        min_col, min_row, max_col, max_row = range_boundaries(f"{tl}:{br}")
        out = []
        for r in range(min_row, max_row+1):
            row = []
            for c in range(min_col, max_col+1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, int):
                    row.append(val & 0xFF)
                else:
                    row.append(int(str(val).strip(), 16) & 0xFF)
            out.append(row)
        return out
    
    cipherkey = read_block("P65", "S68")
    plaintext = read_block("AD104", "AG107")
    return AESInput(cipherkey=cipherkey, plaintext=plaintext)

def parse_hex16(s: str) -> List[int]:
    """Parse 16 hex bytes from string"""
    toks = s.replace(",", " ").split()
    if len(toks) != 16:
        raise ValueError("Expected 16 hex bytes")
    return [int(t, 16) & 0xFF for t in toks]

def list16_to_matrix4x4_columnmajor(vals: List[int]) -> List[List[int]]:
    """Convert 16 bytes to 4x4 matrix in column-major order (AES standard)"""
    matrix = [[0]*4 for _ in range(4)]
    idx = 0
    for col in range(4):
        for row in range(4):
            matrix[row][col] = vals[idx]
            idx += 1
    return matrix

def main():
    ap = argparse.ArgumentParser(description="AES-128 Encryption → Excel Detail Logger")
    ap.add_argument("--src", type=str, default=None, help="Path to source workbook (reads CipherKey dan Plaintext)")
    ap.add_argument("--out", type=str, required=True, help="Output .xlsx path")
    ap.add_argument("--cipherkey-hex", type=str, default=None, help="16 hex bytes untuk cipher key (COLUMN-MAJOR)")
    ap.add_argument("--plaintext-hex", type=str, default=None, help="16 hex bytes untuk plaintext (COLUMN-MAJOR)")
    args = ap.parse_args()

    if args.src:
        aes_in = read_from_source_excel(Path(args.src))
    else:
        if not (args.cipherkey_hex and args.plaintext_hex):
            raise SystemExit("If --src is not provided, you must pass --cipherkey-hex and --plaintext-hex (16 bytes each).")
        ck = list16_to_matrix4x4_columnmajor(parse_hex16(args.cipherkey_hex))
        pt = list16_to_matrix4x4_columnmajor(parse_hex16(args.plaintext_hex))
        aes_in = AESInput(cipherkey=ck, plaintext=pt)

    out_path = build_workbook(aes_in, Path(args.out))
    print(f"Saved workbook to: {out_path}")

if __name__ == "__main__":
    main()
